"""Write the Morning Prep Dashboard (xlsx) from tomorrow's recommendations + backtest KPIs."""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from config import HERE

NAVY="1F3864"; BLUE="2E6CDF"; GREEN="2E8B57"; GOLD="E8A33D"; INK="1F2421"; GREY="8A8079"; WHITE="FFFFFF"
def F(sz=10,b=False,c=INK): return Font(name="Calibri",size=sz,bold=b,color=c)
def fl(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color="D9D2C7"); bd=Border(thin,thin,thin,thin)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True); lft=Alignment("left",vertical="center",wrap_text=True)


def write(recs, summary, target_dow, out=None):
    out = out or os.path.join(HERE, "morning_report.xlsx")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Morning_Prep"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "SUKIMART · MORNING PREP DASHBOARD"; ws["A1"].font=F(15,True,WHITE); ws["A1"].fill=fl(NAVY); ws.merge_cells("A1:F1")
    ws["A2"] = f"Tomorrow is {target_dow}. Prepare this much — and no more.  (system recommends · Tina decides)"
    ws["A2"].font=F(9,False,GREY); ws.merge_cells("A2:F2")
    cards=[("Items to prep",str(len(recs)),BLUE),
           ("₱ saved / month",f"{summary['uplift_month']:,.0f}",GREEN),
           ("Spoilage",f"{summary['spoil_gut']*100:.0f}% → {summary['spoil_sys']*100:.0f}%",GOLD)]
    for i,(lab,val,col) in enumerate(cards):
        c0=1+i*2
        lc=ws.cell(4,c0,lab); lc.font=F(10,True,WHITE); lc.fill=fl(col); lc.alignment=ctr; ws.merge_cells(start_row=4,start_column=c0,end_row=4,end_column=c0+1)
        vc=ws.cell(5,c0,val); vc.font=F(20,True,col); vc.alignment=ctr; ws.merge_cells(start_row=5,start_column=c0,end_row=6,end_column=c0+1)
        for rr in(4,5,6):
            for cc in(c0,c0+1): ws.cell(rr,cc).border=bd
    ws.cell(8,1,"TODAY'S PREP LIST").font=F(11,True,NAVY)
    for c,h in enumerate(["Item","PREPARE","Forecast","Method","Why"],1):
        cell=ws.cell(9,c,h); cell.font=F(10,True,WHITE); cell.fill=fl(BLUE); cell.alignment=ctr; cell.border=bd
    for k,r in enumerate(recs):
        row=10+k
        for c,v in enumerate([r["name"],r["prep"],r["forecast"],r["method"],r["why"]],1):
            cell=ws.cell(row,c,v); cell.border=bd; cell.alignment=lft if c in(1,4,5) else ctr
        ws.cell(row,1).font=F(10,True); ws.cell(row,2).font=F(12,True,GREEN)
    ws.cell(10+len(recs)+1,1,'"Cook to the forecast, not to a guess."  ·  '
            f"backtest: WAPE {summary['wape']*100:.0f}% · service {summary['service_level']*100:.0f}%").font=F(10,True,NAVY)
    for i,w in enumerate([18,9,9,14,24],1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"; ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.page_margins.left = ws.page_margins.right = 0.3; ws.page_margins.top = ws.page_margins.bottom = 0.4
    wb.save(out); return out
