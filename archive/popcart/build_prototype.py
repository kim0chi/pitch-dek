"""
PopCart PH — Forecasting & Reorder Control System (prototype builder)
Generates PopCart_Reorder_System.xlsx with LIVE formulas (not static values).
Open in Google Sheets (upload) or Excel; formulas recalc on open.

Demonstrates Slide 9: ingest -> forecast (weighted 7/14/28d + viral multiplier)
-> reorder point -> alert -> human decides.

Run: .venv/bin/python build_prototype.py
NOTE: daily sales + TikTok index are ILLUSTRATIVE sample inputs (case allows
assumptions). The LOGIC is the deliverable, not the sample numbers.
"""
import random
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

random.seed(42)
TODAY = date(2026, 6, 21)
N = 28  # days of history

# name, base/day, price, lead_time(d), stock_on_hand, trend, tiktok_index(0-100)
SKUS = [
    ("Freeze-Dried Candy Mix",        300, 180, 14, 2500, "viral_up",  88),
    ("Spicy Korean Ramen Bundles",    240, 320, 30, 5000, "steady",    64),
    ("Giant Bubble Gum Tubs",         180, 220,  7, 3200, "cooling",   45),
    ("Assorted Chocolate Crunch Bars",160, 150, 21, 3000, "steady_up", 70),
    ("Sour Belts & Bulk Packs",       120, 120,  7, 2600, "steady",    50),
]

def gen_series(base, trend):
    out = []
    for i in range(N):
        v = base * random.uniform(0.88, 1.12)
        dfe = N - 1 - i               # 0 == today, counting back
        if dfe < 8:
            t = (8 - dfe) / 8         # ramp 0->1 toward today
            if trend == "viral_up":   v *= 1 + 0.35 * t
            elif trend == "cooling":  v *= 1 - 0.20 * t
            elif trend == "steady_up":v *= 1 + 0.12 * t
        out.append(int(round(v)))
    return out

series = {name: gen_series(base, trend) for (name, base, _p, _l, _s, trend, _tk) in SKUS}

# ----- styles --------------------------------------------------------------
NAVY   = "1F3864"; BLUE = "2E5496"; LIGHT = "D9E1F2"; GREY = "F2F2F2"
RED    = "FFC7CE"; REDF = "9C0006"; GREEN = "C6EFCE"; GREENF = "006100"
GOLD   = "FFF2CC"
def font(sz=10, b=False, color="000000"): return Font(size=sz, bold=b, color=color, name="Calibri")
def fill(c): return PatternFill("solid", fgColor=c)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",  vertical="center", wrap_text=True)
right  = Alignment(horizontal="right", vertical="center")

wb = openpyxl.Workbook()

# ===========================================================================
# SHEET 1: Daily_Sales  (INGEST)
# ===========================================================================
ws = wb.active; ws.title = "Daily_Sales"
ws["A1"] = "STEP 1 — UPDATE: Daily units sold (scheduled export first; approved connector later)"
ws["A1"].font = font(13, True, "FFFFFF"); ws["A1"].fill = fill(NAVY)
ws.merge_cells("A1:F1"); ws["A1"].alignment = left
ws["A2"] = "ILLUSTRATIVE 28-day history. Case-based product speeds are used as the baseline; this is not live PopCart data."
ws["A2"].font = font(9, False, "808080"); ws.merge_cells("A2:F2")

hdr = ["Date"] + [s[0] for s in SKUS]
for c, h in enumerate(hdr, 1):
    cell = ws.cell(3, c, h); cell.font = font(10, True, "FFFFFF"); cell.fill = fill(BLUE)
    cell.alignment = center; cell.border = border
for i in range(N):                              # rows 4..31
    r = 4 + i
    d = TODAY - timedelta(days=(N - 1 - i))
    dc = ws.cell(r, 1, d.strftime("%b %d")); dc.alignment = center; dc.border = border
    if i == N - 1: dc.value = d.strftime("%b %d") + " (today)"; dc.font = font(10, True)
    for k, s in enumerate(SKUS):
        cell = ws.cell(r, 2 + k, series[s[0]][i]); cell.alignment = right; cell.border = border
        if i >= N - 7: cell.fill = fill(GREY)   # shade last 7 days
ws.column_dimensions["A"].width = 14
for k in range(len(SKUS)): ws.column_dimensions[get_column_letter(2 + k)].width = 16
ws.freeze_panes = "B4"

# ===========================================================================
# SHEET 2: Control_Panel  (FORECAST + COMPUTE + ALERT)  <-- hero screenshot
# ===========================================================================
cp = wb.create_sheet("Control_Panel")
cp["A1"] = "STEP 2 — REORDER CONTROL PANEL  ·  \"the fuel light for every SKU\""
cp["A1"].font = font(13, True, "FFFFFF"); cp["A1"].fill = fill(NAVY)
cp.merge_cells("A1:P1"); cp["A1"].alignment = left
cp["A2"] = ("Reorder Point = forecast daily demand × lead time + safety stock. "
            "Status flags 🔴 when stock on hand falls to/below it. System recommends; founder decides.")
cp["A2"].font = font(9, False, "808080"); cp.merge_cells("A2:P2")

cols = ["SKU", "Price ₱", "Lead\n(days)", "Stock\non hand",
        "Avg/day\n(7d)", "Avg/day\n(14d)", "Avg/day\n(28d)", "TikTok\nidx (0-100)",
        "Viral\n×", "Forecast\n/day", "Safety\nstock", "Reorder\npoint",
        "STATUS", "Suggested\norder qty", "Days\ncover", "₱ contribution\nif OOS/day"]
for c, h in enumerate(cols, 1):
    cell = cp.cell(3, c, h); cell.font = font(9, True, "FFFFFF"); cell.fill = fill(BLUE)
    cell.alignment = center; cell.border = border

dcol = {name: get_column_letter(2 + k) for k, name in enumerate(series)}  # Daily_Sales col per SKU
for k, s in enumerate(SKUS):
    r = 4 + k
    name, base, price, lead, stock, trend, tk = s
    L = dcol[name]
    ds = lambda a, b: f"Daily_Sales!{L}{a}:{L}{b}"
    cp.cell(r, 1, name)
    cp.cell(r, 2, price)
    cp.cell(r, 3, lead)
    cp.cell(r, 4, stock)
    cp.cell(r, 5, f"=ROUND(AVERAGE({ds(25,31)}),0)")            # 7d  (rows 25-31)
    cp.cell(r, 6, f"=ROUND(AVERAGE({ds(18,31)}),0)")            # 14d (rows 18-31)
    cp.cell(r, 7, f"=ROUND(AVERAGE({ds(4,31)}),0)")             # 28d (rows 4-31)
    cp.cell(r, 8, tk)                                           # TikTok index (input)
    cp.cell(r, 9, f"=1+MIN(0.4,MAX(0,(H{r}-60)/100))")         # viral multiplier
    cp.cell(r,10, f"=ROUND((0.5*E{r}+0.3*F{r}+0.2*G{r})*I{r},0)")  # forecast/day
    cp.cell(r,11, f"=ROUND(1.65*STDEV({ds(4,31)})*SQRT(C{r}),0)")  # safety stock (Z=1.65 ~95%)
    cp.cell(r,12, f"=ROUND(J{r}*C{r}+K{r},0)")                  # reorder point
    cp.cell(r,13, f'=IF(D{r}<=L{r},"🔴 REORDER NOW","🟢 OK")')  # status
    cp.cell(r,14, f"=IF(D{r}<=L{r},ROUNDUP(J{r}*(C{r}+14)+K{r}-D{r},0),0)")  # order qty
    cp.cell(r,15, f"=ROUND(D{r}/J{r},1)")                       # days of cover
    cp.cell(r,16, f"=ROUND(J{r}*B{r}*0.3,0)")                   # profit risk/day (30% contribution)
    for c in range(1, 17):
        cell = cp.cell(r, c); cell.border = border
        cell.alignment = left if c == 1 else center
        if c in (2, 16): cell.number_format = "#,##0"
        if c == 9: cell.number_format = "0.00"
        if c == 15: cell.number_format = "0.0"
    cp.cell(r, 1).font = font(10, True)
    cp.cell(r, 13).font = font(10, True)
    cp.cell(r, 16).font = font(10, True)

# conditional formatting on STATUS column (M4:M8)
cp.conditional_formatting.add("M4:M8",
    FormulaRule(formula=['ISNUMBER(SEARCH("REORDER",M4))'], fill=fill(RED), font=font(10, True, REDF)))
cp.conditional_formatting.add("M4:M8",
    FormulaRule(formula=['ISNUMBER(SEARCH("OK",M4))'], fill=fill(GREEN), font=font(10, True, GREENF)))

widths = [30,9,7,8,8,8,8,9,7,9,8,9,16,11,7,11]
for i, w in enumerate(widths, 1): cp.column_dimensions[get_column_letter(i)].width = w
cp.row_dimensions[3].height = 30
cp.freeze_panes = "B4"

# legend
lr = 11
cp.cell(lr,1,"How to read this panel:").font = font(10, True, NAVY)
notes = [
 "• Forecast/day = weighted recency (50% of 7-day, 30% of 14-day, 20% of 28-day) × viral multiplier.",
 "• Viral × rises with the TikTok engagement index (leading signal) — pre-empts a spike before sales fully show it.",
 "• Safety stock = 1.65 × demand volatility (st-dev) × √lead time  →  ~95% service level; bigger when sales are choppy or lead times long.",
 "• Reorder point (the fuel light) = forecast/day × lead time + safety stock. Order BEFORE stock hits it, because restock takes the lead time to arrive.",
 "• ₱ contribution if OOS/day = forecast/day × price × 30% — an illustrative exposure if the item becomes unavailable, not a claim of today's actual loss.",
]
for i, t in enumerate(notes):
    cp.cell(lr+1+i, 1, t).font = font(9); cp.merge_cells(start_row=lr+1+i, start_column=1, end_row=lr+1+i, end_column=16)

# ===========================================================================
# SHEET 3: Dashboard  (the founder's morning view)
# ===========================================================================
db = wb.create_sheet("Dashboard")
db["A1"] = "PopCart — Morning Control Dashboard"
db["A1"].font = font(15, True, "FFFFFF"); db["A1"].fill = fill(NAVY)
db.merge_cells("A1:F1"); db["A1"].alignment = left
db["A2"] = "Prototype using case baselines + illustrative operating data. All figures live-linked to the Control Panel."
db["A2"].font = font(9, False, "808080"); db.merge_cells("A2:F2")

# KPI cards
cards = [
    ("SKUs to reorder TODAY", '=COUNTIF(Control_Panel!M4:M8,"*REORDER*")', NAVY),
    ("₱ contribution exposed if OOS/day", '=SUMIF(Control_Panel!M4:M8,"*REORDER*",Control_Panel!P4:P8)', REDF),
    ("Viral items rising 🔥", '=COUNTIF(Control_Panel!I4:I8,">1.1")', "C55A11"),
]
for i, (lab, fla, col) in enumerate(cards):
    c0 = 1 + i * 2
    lc = db.cell(4, c0, lab); lc.font = font(10, True, "FFFFFF"); lc.fill = fill(col)
    lc.alignment = center; db.merge_cells(start_row=4, start_column=c0, end_row=4, end_column=c0+1)
    vc = db.cell(5, c0, fla); vc.font = font(22, True, col); vc.fill = fill(GREY)
    vc.alignment = center; db.merge_cells(start_row=5, start_column=c0, end_row=6, end_column=c0+1)
    if i == 1: vc.number_format = "#,##0"
    for rr in (4,5,6):
        for cc in (c0, c0+1):
            db.cell(rr, cc).border = border

# action table
db.cell(8,1,"TODAY'S ACTION LIST  (system recommends — founder approves)").font = font(11, True, NAVY)
ah = ["SKU","Status","Suggested order qty","Days cover left","₱ if OOS/day","Trend"]
for c,h in enumerate(ah,1):
    cell = db.cell(9,c,h); cell.font = font(10, True, "FFFFFF"); cell.fill = fill(BLUE)
    cell.alignment = center; cell.border = border
for k in range(5):
    r = 10 + k; cpr = 4 + k
    refs = [f"=Control_Panel!A{cpr}", f"=Control_Panel!M{cpr}", f"=Control_Panel!N{cpr}",
            f"=Control_Panel!O{cpr}", f"=Control_Panel!P{cpr}",
            f'=IF(Control_Panel!I{cpr}>1.1,"🔥 rising","steady")']
    for c, f in enumerate(refs, 1):
        cell = db.cell(r, c, f); cell.border = border
        cell.alignment = left if c == 1 else center
        if c == 5: cell.number_format = "#,##0"
    db.cell(r,1).font = font(10, True)
db.conditional_formatting.add("B10:B14",
    FormulaRule(formula=['ISNUMBER(SEARCH("REORDER",B10))'], fill=fill(RED), font=font(10, True, REDF)))
db.conditional_formatting.add("B10:B14",
    FormulaRule(formula=['ISNUMBER(SEARCH("OK",B10))'], fill=fill(GREEN), font=font(10, True, GREENF)))
db.cell(16,1,'"Reorder before the shelf is empty — not after. The fuel light, for snacks."').font = font(10, True, NAVY)
for i,w in enumerate([30,16,18,15,12,12],1): db.column_dimensions[get_column_letter(i)].width = w

# ===========================================================================
# SHEET 4: README
# ===========================================================================
rd = wb.create_sheet("README")
rd.column_dimensions["A"].width = 110
lines = [
 ("PopCart PH — Forecasting & Reorder Control System (prototype)", 14, True, NAVY),
 ("Built for LAMBO 2026, Slide 9. A working demo of the proposed decision process — not PopCart's finished or live system.", 10, False, "808080"),
 ("", 10, False, "000000"),
 ("THE 5-STEP WORKFLOW", 12, True, BLUE),
 ("1. UPDATE  — Start with scheduled exports into Excel/Power Query, 2–3×/day. Use an approved connector later if PopCart chooses.", 10, False, "000000"),
 ("           Replaces the manual twice-a-day Excel — like switching from checking your bank balance twice a day to online banking.", 10, False, "808080"),
 ("2. COMPARE — estimate recent demand, then compare days of stock with supplier delivery time.", 10, False, "000000"),
 ("3. WARN     — flag products that may run out before the next delivery can arrive.", 10, False, "000000"),
 ("4. DECIDE   — founder checks cash, supplier availability, and minimum order size before approving.", 10, False, "000000"),
 ("5. LEARN    — compare alerts with actual results weekly; add trend signals only after the basic data is reliable.", 10, False, "000000"),
 ("", 10, False, "000000"),
 ("WHY IT FIXES THE PHP 7.35M LEAK", 12, True, BLUE),
 ("Today PopCart reorders on gut feel and only sees empty shelves AFTER they happen — fatal when restock takes 14–30 days.", 10, False, "000000"),
 ("The reorder point is the low-fuel light: it warns with enough road left to refill before the tank hits empty.", 10, False, "808080"),
 ("Note the long-lead items (ramen 30d, choco 21d) trigger earliest — exactly where blind reordering hurts most.", 10, False, "000000"),
 ("", 10, False, "000000"),
 ("HONESTY / ASSUMPTIONS", 12, True, BLUE),
 ("• CASE-BASED: product velocity, price, and supplier lead time. ILLUSTRATIVE: daily history, current stock, and trend index.", 10, False, "000000"),
 ("• Profit-at-risk uses a conservative 30% CONTRIBUTION margin (recovered sales still pay platform fees + shipping), not 50% gross.", 10, False, "000000"),
 ("• Z=1.65 (~95% service level). Weights, safety days, and the viral cap are tunable weekly.", 10, False, "000000"),
 ("", 10, False, "000000"),
 ("AI-USE DISCLOSURE", 12, True, BLUE),
 ("Proposed starting tools: Excel + Power Query + a barcode tool, building on PopCart's current Excel process. No custom app required.", 10, False, "000000"),
 ("This prototype's formulas + sample data were drafted with AI assistance and verified by the team against the case exhibits.", 10, False, "000000"),
]
for i, (t, sz, b, col) in enumerate(lines, 1):
    c = rd.cell(i, 1, t); c.font = font(sz, b, col); c.alignment = left

# print setup: fit each sheet to one page wide, landscape, for clean PDF/screenshots
for sheet in (ws, cp, db):
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True if sheet.sheet_properties.pageSetUpPr else None
from openpyxl.worksheet.properties import PageSetupProperties
for sheet in (ws, cp, db, rd):
    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_margins.left = sheet.page_margins.right = 0.3
    sheet.page_margins.top = sheet.page_margins.bottom = 0.4

wb.save("PopCart_Reorder_System.xlsx")
print("Saved PopCart_Reorder_System.xlsx  (4 tabs, live formulas)")

# ---------------------------------------------------------------------------
# VERIFICATION: replicate the formula logic in Python to prove what the
# spreadsheet will display (and confirm the red/green alert mix).
# ---------------------------------------------------------------------------
import statistics
print("\nEXPECTED OUTPUT (what the live formulas will compute):")
print(f"{'SKU':30s}{'fcast/d':>8}{'SS':>6}{'ROP':>7}{'stock':>7}  STATUS        {'ordQty':>7}{'₱risk/d':>9}")
total_risk = 0; reorder = 0; viral = 0
for (name, base, price, lead, stock, trend, tk) in SKUS:
    s = series[name]
    a7, a14, a28 = round(statistics.mean(s[-7:])), round(statistics.mean(s[-14:])), round(statistics.mean(s))
    vmult = 1 + min(0.4, max(0, (tk - 60) / 100))
    fcast = round((0.5*a7 + 0.3*a14 + 0.2*a28) * vmult)
    ss = round(1.65 * statistics.stdev(s) * (lead ** 0.5))
    rop = round(fcast * lead + ss)
    status = "🔴 REORDER NOW" if stock <= rop else "🟢 OK"
    qty = max(0, round(fcast*(lead+14) + ss - stock)) if stock <= rop else 0
    risk = round(fcast * price * 0.30)
    if stock <= rop: reorder += 1; total_risk += risk
    if vmult > 1.1: viral += 1
    print(f"{name:30s}{fcast:>8}{ss:>6}{rop:>7}{stock:>7}  {status:13s}{qty:>7}{risk:>9,}")
print(f"\nDashboard KPIs -> reorder today: {reorder} | viral rising: {viral} | ₱ at risk/day: {total_risk:,}")
